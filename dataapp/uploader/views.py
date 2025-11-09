import os
from pathlib import Path
from io import BytesIO, StringIO

from django.shortcuts import render, redirect
from django.conf import settings
from django.http import HttpResponse, FileResponse, Http404

import pandas as pd
import arff
from sklearn.model_selection import train_test_split

# plotting
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

# PDF & Excel
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from django.utils.text import slugify

MEDIA_ROOT = Path(settings.MEDIA_ROOT) if hasattr(settings, "MEDIA_ROOT") else Path("media")
UPLOADS_DIR = MEDIA_ROOT / "uploads"
SPLITS_DIR = MEDIA_ROOT / "splits"
PLOTS_DIR = MEDIA_ROOT / "plots"

for d in (UPLOADS_DIR, SPLITS_DIR, PLOTS_DIR):
    d.mkdir(parents=True, exist_ok=True)


# --- helpers ---
def load_arff(path):
    with open(path, "r") as f:
        dataset = arff.load(f)
        attributes = [a[0] for a in dataset["attributes"]]
        return pd.DataFrame(dataset["data"], columns=attributes)


def df_info_text(df: pd.DataFrame) -> str:
    buf = StringIO()
    df.info(buf=buf)
    return buf.getvalue()


def save_dataframe_csv(df: pd.DataFrame, path: Path):
    df.to_csv(path, index=False)


def save_dataframe_arff(df: pd.DataFrame, path: Path, relation: str = "dataset"):
    attributes = []
    for col in df.columns:
        dtype = df[col].dtype
        if pd.api.types.is_numeric_dtype(dtype):
            typ = 'NUMERIC'
        else:
            # For ARFF liac-arff expects list of nominal values or 'STRING'/'NUMERIC', we'll use STRING
            typ = 'STRING'
        attributes.append((col, typ))
    arff_dict = {
        'description': '',
        'relation': relation,
        'attributes': attributes,
        'data': df.values.tolist()
    }
    with open(path, 'w') as f:
        arff.dump(arff_dict, f)


def make_histogram(df: pd.DataFrame, base_slug: str) -> str:
    """Return relative path to saved histogram PNG."""
    # choose first numeric column
    num_cols = df.select_dtypes(include='number').columns.tolist()
    if not num_cols:
        # fallback: use first column
        col = df.columns[0]
    else:
        col = num_cols[0]

    plt.figure(figsize=(6, 3.5))
    sns.histplot(df[col].dropna(), kde=True, color="#0ea5ff")
    plt.title(f"Distribución de {col}", color="white")
    plt.tight_layout()
    p = PLOTS_DIR / f"{base_slug}_hist.png"
    plt.savefig(p, dpi=120, bbox_inches="tight", facecolor="#0b0b0c")
    plt.close()
    return str(p.relative_to(MEDIA_ROOT))


def make_heatmap_train(train_df: pd.DataFrame, base_slug: str) -> str:
    # compute correlation on numeric columns
    df_num = train_df.select_dtypes(include='number')
    if df_num.shape[1] < 2:
        # not enough numeric columns; create an empty placeholder image
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.text(0.5, 0.5, "No hay suficientes columnas numéricas para correlación",
                ha='center', va='center', color='white')
        ax.axis('off')
        p = PLOTS_DIR / f"{base_slug}_heatmap.png"
        fig.savefig(p, dpi=120, bbox_inches="tight", facecolor="#0b0b0c")
        plt.close(fig)
        return str(p.relative_to(MEDIA_ROOT))

    corr = df_num.corr()
    plt.figure(figsize=(7, 6))
    sns.heatmap(corr, annot=False, cmap='Blues', cbar=True)
    plt.title("Matriz de correlación (train)", color="white")
    plt.tight_layout()
    p = PLOTS_DIR / f"{base_slug}_heatmap.png"
    plt.savefig(p, dpi=120, bbox_inches="tight", facecolor="#0b0b0c")
    plt.close()
    return str(p.relative_to(MEDIA_ROOT))


# --- views ---
def upload_dataset(request):
    """
    GET: show upload form
    POST: receive file, parse, make histogram, show preview and division controls
    """
    context = {
        "info": None,
        "head": None,
        "hist_path": None,
        "show_split": False,
    }

    if request.method == "POST" and request.FILES.get("dataset"):
        file = request.FILES["dataset"]
        filename = file.name
        lname = filename.lower()
        base_slug = slugify(Path(filename).stem) or "dataset"

        # save uploaded file
        UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
        save_path = UPLOADS_DIR / filename
        with open(save_path, "wb+") as dest:
            for chunk in file.chunks():
                dest.write(chunk)

        # load df
        try:
            if lname.endswith(".csv"):
                df = pd.read_csv(save_path)
            elif lname.endswith(".xlsx") or lname.endswith(".xls"):
                df = pd.read_excel(save_path)
            elif lname.endswith(".arff"):
                df = load_arff(save_path)
            else:
                context["info"] = "Formato no soportado (usa CSV, XLSX o ARFF)."
                return render(request, "uploader/upload.html", context)
        except Exception as e:
            context["info"] = f"Error leyendo archivo: {e}"
            return render(request, "uploader/upload.html", context)

        # prepare outputs
        context["info"] = df_info_text(df)
        context["head"] = df.head().to_html(classes="table table-sm", index=False)
        hist_rel = make_histogram(df, base_slug)
        context["hist_path"] = f"{settings.MEDIA_URL}{hist_rel}"
        context["show_split"] = True

        # store info in session so split view can find the uploaded file and slug
        request.session["uploaded_filename"] = filename
        request.session["dataset_slug"] = base_slug

    return render(request, "uploader/upload.html", context)


def split_dataset(request):
    """
    POST: performs splitting using the uploaded file referenced in session.
    Renders same template with splits preview, heatmap and download options.
    """
    if request.method != "POST":
        return redirect("upload")

    filename = request.session.get("uploaded_filename")
    base_slug = request.session.get("dataset_slug")
    if not filename or not base_slug:
        return redirect("upload")

    save_path = UPLOADS_DIR / filename
    if not save_path.exists():
        return redirect("upload")

    lname = filename.lower()
    try:
        if lname.endswith(".csv"):
            df = pd.read_csv(save_path)
        elif lname.endswith(".xlsx") or lname.endswith(".xls"):
            df = pd.read_excel(save_path)
        elif lname.endswith(".arff"):
            df = load_arff(save_path)
        else:
            return redirect("upload")
    except Exception:
        return redirect("upload")

    # get options from form
    include_val = request.POST.get("include_val") == "on"
    save_splits = request.POST.get("save_splits") == "on"
    fmt = request.POST.get("format", "csv").lower()

    try:
        train_pct = float(request.POST.get("train_pct", 70))
        if include_val:
            val_pct = float(request.POST.get("val_pct", 10))
            test_pct = float(request.POST.get("test_pct", 20))
        else:
            val_pct = 0.0
            test_pct = float(request.POST.get("test_pct", 30))
    except Exception:
        return redirect("upload")

    # validate sum
    total_pct = train_pct + val_pct + test_pct
    if abs(total_pct - 100.0) > 1e-6:
        context = {"info": "Los porcentajes deben sumar 100.", "show_split": True}
        return render(request, "uploader/upload.html", context)

    # perform split
    if include_val:
        train_ratio = train_pct / 100.0
        rest_ratio = 1.0 - train_ratio
        val_ratio_within_rest = val_pct / (val_pct + test_pct) if (val_pct + test_pct) > 0 else 0.0
        train_df, rest_df = train_test_split(df, train_size=train_ratio, random_state=42, shuffle=True)
        # rest_df split into val and test
        if val_pct > 0:
            val_df, test_df = train_test_split(rest_df, test_size=(1 - val_ratio_within_rest), random_state=42, shuffle=True)
        else:
            val_df = pd.DataFrame()
            test_df = rest_df
    else:
        train_ratio = train_pct / 100.0
        train_df, test_df = train_test_split(df, train_size=train_ratio, random_state=42, shuffle=True)
        val_df = pd.DataFrame()

    # prepare previews and heatmap
    context = {
        "info": df_info_text(df),
        "train_preview": train_df.head().to_html(classes="table table-sm", index=False),
        "test_preview": test_df.head().to_html(classes="table table-sm", index=False),
        "val_preview": val_df.head().to_html(classes="table table-sm", index=False) if not val_df.empty else None,
        "splits_info": {
            "total_rows": len(df),
            "train_rows": len(train_df),
            "val_rows": len(val_df) if not val_df.empty else 0,
            "test_rows": len(test_df),
        },
        "saved_msg": None,
        "hist_path": None,
        "heatmap_path": None,
    }

    # generate heatmap for train
    heat_rel = make_heatmap_train(train_df, base_slug)
    context["heatmap_path"] = f"{settings.MEDIA_URL}{heat_rel}"

    # also keep the previously generated histogram (if exists)
    hist_file = PLOTS_DIR / f"{base_slug}_hist.png"
    if hist_file.exists():
        context["hist_path"] = f"{settings.MEDIA_URL}{hist_file.relative_to(MEDIA_ROOT)}"

    # saving splits if requested
    if save_splits:
        SPLITS_DIR.mkdir(parents=True, exist_ok=True)
        try:
            if fmt == "csv":
                save_dataframe_csv(train_df, SPLITS_DIR / f"{base_slug}_train.csv")
                save_dataframe_csv(test_df, SPLITS_DIR / f"{base_slug}_test.csv")
                if not val_df.empty:
                    save_dataframe_csv(val_df, SPLITS_DIR / f"{base_slug}_val.csv")
            elif fmt == "arff":
                save_dataframe_arff(train_df, SPLITS_DIR / f"{base_slug}_train.arff", relation=f"{base_slug}_train")
                save_dataframe_arff(test_df, SPLITS_DIR / f"{base_slug}_test.arff", relation=f"{base_slug}_test")
                if not val_df.empty:
                    save_dataframe_arff(val_df, SPLITS_DIR / f"{base_slug}_val.arff", relation=f"{base_slug}_val")
            context["saved_msg"] = f"Archivos guardados en media/splits/ ({fmt.upper()})"
        except Exception as e:
            context["saved_msg"] = f"Error guardando splits: {e}"

    # store last_split slug so download views know which files to use
    request.session["last_split_slug"] = base_slug
    # also save a simple mapping of filenames to allow download construction
    request.session["last_split_format"] = fmt

    return render(request, "uploader/upload.html", context)


def download_pdf(request):
    """
    Build a PDF 'reporte de laboratorio' containing:
    - Title "Análisis del dataset"
    - Info text
    - The two images (histogram + heatmap) if present
    - Short tables (train/test/val heads)
    """
    base_slug = request.session.get("last_split_slug")
    if not base_slug:
        raise Http404("No hay análisis previo para generar PDF.")

    # Source files
    hist_p = PLOTS_DIR / f"{base_slug}_hist.png"
    heat_p = PLOTS_DIR / f"{base_slug}_heatmap.png"
    splits = {
        "train": SPLITS_DIR / f"{base_slug}_train.csv",
        "test": SPLITS_DIR / f"{base_slug}_test.csv",
        "val": SPLITS_DIR / f"{base_slug}_val.csv"
    }

    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Análisis del dataset", styles['Title']))
    story.append(Spacer(1, 12))

    # info: try to open original upload info if exists
    upload_fname = request.session.get("uploaded_filename")
    if upload_fname:
        story.append(Paragraph(f"Archivo: {upload_fname}", styles['Normal']))
        story.append(Spacer(1, 8))

    # Images
    if hist_p.exists():
        story.append(Paragraph("Histograma (columna numérica detectada)", styles['Heading3']))
        story.append(Spacer(1, 4))
        story.append(RLImage(str(hist_p), width=400, height=200))
        story.append(Spacer(1, 8))

    if heat_p.exists():
        story.append(Paragraph("Heatmap de correlación (train)", styles['Heading3']))
        story.append(Spacer(1, 4))
        story.append(RLImage(str(heat_p), width=420, height=300))
        story.append(Spacer(1, 8))

    # Include small samples of CSV splits if exist (read first 5 rows)
    def add_table_from_csv(path, title):
        if path.exists():
            df = pd.read_csv(path)
            story.append(Paragraph(title, styles['Heading4']))
            story.append(Spacer(1, 4))
            # Convert to small table (head)
            head = df.head().astype(str)
            data = [list(head.columns)]
            for _, row in head.iterrows():
                data.append(list(row.values))
            tbl = Table(data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b2b2b')),
                ('TEXTCOLOR', (0,0),(-1,0), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 12))

    add_table_from_csv(splits['train'], "Train — primeras filas")
    add_table_from_csv(splits['val'], "Validation — primeras filas")
    add_table_from_csv(splits['test'], "Test — primeras filas")

    doc.build(story)
    buffer.seek(0)

    response = FileResponse(buffer, as_attachment=True, filename=f"{base_slug}_report.pdf")
    return response


def download_excel(request):
    """
    Build an Excel file with sheets: info, train, val, test, and insert images.
    """
    base_slug = request.session.get("last_split_slug")
    if not base_slug:
        raise Http404("No hay análisis previo para generar Excel.")

    hist_p = PLOTS_DIR / f"{base_slug}_hist.png"
    heat_p = PLOTS_DIR / f"{base_slug}_heatmap.png"
    splits = {
        "train": SPLITS_DIR / f"{base_slug}_train.csv",
        "test": SPLITS_DIR / f"{base_slug}_test.csv",
        "val": SPLITS_DIR / f"{base_slug}_val.csv"
    }

    wb = Workbook()
    # Info sheet
    ws_info = wb.active
    ws_info.title = "info"
    upload_fname = request.session.get("uploaded_filename", "")
    ws_info.append(["Archivo", upload_fname])
    ws_info.append(["Título", "Análisis del dataset"])
    # Try to include df.info() text
    # If original uploaded file exists, try to compute info
    uploaded = UPLOADS_DIR / upload_fname if upload_fname else None
    if uploaded and uploaded.exists():
        try:
            if str(uploaded).lower().endswith(".arff"):
                df = load_arff(uploaded)
            else:
                df = pd.read_csv(uploaded) if str(uploaded).lower().endswith(".csv") else pd.read_excel(uploaded)
            info_text = df_info_text(df)
            for line in info_text.splitlines():
                ws_info.append([line])
        except Exception:
            pass

    # Add splits as sheets
    for name in ("train", "val", "test"):
        path = splits[name]
        if path.exists():
            df = pd.read_csv(path)
            ws = wb.create_sheet(title=name)
            # write header
            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append(list(row.values))

    # add images sheet
    ws_img = wb.create_sheet(title="graficas")
    row_cursor = 1
    if hist_p.exists():
        img = XLImage(str(hist_p))
        img.anchor = f"A{row_cursor}"
        ws_img.add_image(img)
        row_cursor += 30
    if heat_p.exists():
        img2 = XLImage(str(heat_p))
        img2.anchor = f"A{row_cursor}"
        ws_img.add_image(img2)

    # Save to memory
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return FileResponse(bio, as_attachment=True, filename=f"{base_slug}_analysis.xlsx")
